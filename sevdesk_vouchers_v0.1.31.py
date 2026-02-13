import requests
import argparse
from datetime import datetime
import sys
import urllib3
import os
import shutil
import re
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Versioning
__version__ = "0.1.31"

# UTF-8 Encoding für Windows-Konsole erzwingen
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

def parse_date(date_str):
    """Parst Datum im Format DD.MM.YYYY oder YYYY-MM-DD"""
    # Versuche zuerst DD.MM.YYYY
    try:
        return datetime.strptime(date_str, '%d.%m.%Y')
    except ValueError:
        pass
    
    # Versuche YYYY-MM-DD
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        pass
    
    print(f"Fehler: Datum '{date_str}' muss im Format DD.MM.YYYY oder YYYY-MM-DD sein.")
    sys.exit(1)

def convert_date_to_dir_format(date_str):
    """Konvertiert DD.MM.YYYY zu YYYY.MM.DD für Verzeichnisnamen"""
    dt = parse_date(date_str)
    return dt.strftime('%Y.%m.%d')

def fetch_vouchers(session, voucher_type, verify_ssl=True, debug=False):
    """Ruft Vouchers von der sevDesk API ab.
    
    BUG-FIX v0.1.29: Verbesserte Filterung und Debug-Ausgaben.
    
    Die sevDesk API verwendet:
    - creditDebit=C (Credit) = Eingangsrechnungen (expense) - Du hast etwas GEKAUFT
    - creditDebit=D (Debit) = Ausgangsrechnungen (income) - Du hast etwas VERKAUFT
    
    Args:
        session: requests Session mit Auth-Header
        voucher_type: 'income' (Ausgangsrechnungen) oder 'expense' (Eingangsrechnungen)
        verify_ssl: SSL-Verifizierung
        debug: Debug-Modus
    
    Returns:
        list: Liste von Voucher-Objekten (nur des gewünschten Typs)
    """
    url = "https://my.sevdesk.de/api/v1/Voucher"
    
    # creditDebit-Mapping laut sevDesk API-Dokumentation:
    # 'D' (Debit) = income/Ausgangsrechnungen (Du verkaufst etwas → Geld KOMMT)
    # 'C' (Credit) = expense/Eingangsrechnungen (Du kaufst etwas → Geld GEHT)
    credit_debit = 'D' if voucher_type == 'income' else 'C'
    
    params = {
        "limit": 1000,
        "embed": "taxRule,supplier",
        "creditDebit": credit_debit
    }
    
    # Zeige IMMER die API-Request-Informationen (nicht nur im Debug-Modus)
    print(f"\n[info] API Request URL: {url}")
    print(f"[info] API Parameters:")
    print(f"       - limit: {params['limit']}")
    print(f"       - creditDebit: {credit_debit} ({'income/Ausgangsrechnungen (Debit)' if voucher_type == 'income' else 'expense/Eingangsrechnungen (Credit)'})")
    
    if debug:
        print(f"[debug] Full params dict: {params}")
    
    try:
        response = session.get(url, params=params, verify=verify_ssl, timeout=30)
        
        # Zeige die tatsächlich gesendete URL
        print(f"[info] Actual request URL: {response.url}")
        
        response.raise_for_status()
        vouchers = response.json().get('objects', [])
        
        print(f"[info] API returned {len(vouchers)} vouchers")
        
        # ====== KRITISCHER DEBUG: Zeige creditDebit-Verteilung ======
        if vouchers:
            credit_debit_distribution = {}
            for v in vouchers:
                cd = v.get('creditDebit', 'NONE')
                credit_debit_distribution[cd] = credit_debit_distribution.get(cd, 0) + 1
            
            print(f"\n[info] creditDebit distribution in API response:")
            for cd_val, count in sorted(credit_debit_distribution.items()):
                label = "income/Ausgangsrechnungen" if cd_val == 'D' else "expense/Eingangsrechnungen" if cd_val == 'C' else "UNKNOWN"
                print(f"       - {cd_val}: {count} vouchers ({label})")
            
            # Zeige erste 3 Vouchers als Beispiel
            print(f"\n[info] Sample of first 3 vouchers:")
            for i, v in enumerate(vouchers[:3], 1):
                print(f"       {i}. ID={v.get('id')}, voucherNumber={v.get('voucherNumber')}, "
                      f"creditDebit={v.get('creditDebit')}, voucherDate={v.get('voucherDate')[:10] if v.get('voucherDate') else 'N/A'}")
            
            # ====== WARNUNG wenn falsche Vouchers zurückgegeben werden ======
            expected_cd = credit_debit
            wrong_vouchers = [v for v in vouchers if v.get('creditDebit') != expected_cd]
            if wrong_vouchers:
                print(f"\n[WARNING] ================================================")
                print(f"[WARNING] API returned {len(wrong_vouchers)} vouchers with WRONG creditDebit!")
                print(f"[WARNING] Expected: creditDebit={expected_cd} ({voucher_type})")
                print(f"[WARNING] These vouchers will be FILTERED OUT locally!")
                print(f"[WARNING] ================================================")
                
                # Zeige einige Beispiele der falschen Vouchers
                print(f"[WARNING] Examples of wrong vouchers:")
                for v in wrong_vouchers[:5]:
                    print(f"       - ID={v.get('id')}, creditDebit={v.get('creditDebit')}, voucherNumber={v.get('voucherNumber')}")
        
        return vouchers
        
    except requests.exceptions.HTTPError as e:
        print(f"[ERROR] HTTP Error: {e}")
        print(f"[ERROR] Response: {e.response.text[:500] if e.response else 'No response'}")
        return []
    except Exception as e:
        print(f"[ERROR] Fehler beim Abrufen der Voucher Daten: {e}")
        return []

def fetch_tags_for_all_vouchers(session, vouchers, verify_ssl=True, debug=False):
    """Ruft Tags für alle Vouchers ab über die TagRelation API.

    BUG-FIX v0.1.31: Verwendet GET /TagRelation statt GET /Tag, da
    GET /Tag?objectName=Voucher&objectId={id} die objectId-Filterung
    NICHT korrekt umsetzt. Der Endpoint gibt Tags für ALLE Vouchers
    zurück, nicht nur für den angefragten — dadurch wurden Expense-Vouchers
    fälschlicherweise als bereits getaggt erkannt, wenn zuvor ein
    Income-Export-Tag existierte.

    GET /TagRelation gibt die exakten Zuordnungen (Tag ↔ Voucher) zurück,
    sodass nur tatsächlich getaggte Vouchers erkannt werden.

    Args:
        session: requests Session mit Auth-Header
        vouchers: Liste von Voucher-Objekten (aus fetch_vouchers)
        verify_ssl: SSL-Verifizierung
        debug: Debug-Modus

    Returns:
        dict: Mapping von voucher_id zu Liste von Tag-Namen
    """
    if not vouchers:
        return {}

    total = len(vouchers)

    # Mapping von str(id) → original id (für konsistente Schlüssel)
    id_map = {}
    for v in vouchers:
        vid = v.get('id')
        if vid is not None:
            id_map[str(vid)] = vid

    print(f"\n[info] Fetching tags for {total} vouchers via TagRelation API (v0.1.31)...")

    # Schritt 1: Alle Tags abrufen (für ID → Name Mapping)
    tag_names_by_id = {}
    try:
        tag_url = "https://my.sevdesk.de/api/v1/Tag"
        tag_params = {"limit": 9999}
        if debug:
            print(f"[debug] Fetching all tags: GET {tag_url}")
        resp = session.get(tag_url, params=tag_params, verify=verify_ssl, timeout=30)
        resp.raise_for_status()
        all_tags = resp.json().get('objects', [])
        for tag in all_tags:
            tid = str(tag.get('id', ''))
            tname = tag.get('name', '')
            if tid and tname:
                tag_names_by_id[tid] = tname
        if debug:
            print(f"[debug] Found {len(tag_names_by_id)} tags in system: {list(tag_names_by_id.values())[:10]}")
    except Exception as e:
        print(f"[WARNING] Konnte Tags nicht abrufen: {e}")

    # Schritt 2: Alle TagRelations abrufen (exakte Zuordnungen Tag ↔ Objekt)
    tags_map = {}
    try:
        rel_url = "https://my.sevdesk.de/api/v1/TagRelation"
        rel_params = {"limit": 9999}
        if debug:
            print(f"[debug] Fetching tag relations: GET {rel_url}")
        resp = session.get(rel_url, params=rel_params, verify=verify_ssl, timeout=30)
        resp.raise_for_status()
        relations = resp.json().get('objects', [])
        if debug:
            print(f"[debug] Found {len(relations)} total tag relations in system")

        matched_count = 0
        for rel in relations:
            obj = rel.get('object', {})
            if obj.get('objectName') != 'Voucher':
                continue

            obj_id_str = str(obj.get('id', ''))
            if obj_id_str not in id_map:
                continue

            original_id = id_map[obj_id_str]

            # Tag-Name ermitteln: erst aus Relation, dann aus Tag-Map
            tag_info = rel.get('tag', {})
            tag_name = tag_info.get('name', '')
            if not tag_name:
                tag_id = str(tag_info.get('id', ''))
                tag_name = tag_names_by_id.get(tag_id, f'Tag#{tag_id}')

            if original_id not in tags_map:
                tags_map[original_id] = []
            if tag_name and tag_name not in tags_map[original_id]:
                tags_map[original_id].append(tag_name)
                matched_count += 1

        if debug:
            print(f"[debug] Matched {matched_count} tag relations to requested vouchers")

    except Exception as e:
        print(f"[WARNING] Konnte TagRelations nicht abrufen: {e}")
        print(f"[WARNING] Alle Vouchers werden als ungetaggt behandelt")

    # Fehlende Voucher-IDs mit leeren Listen auffüllen
    for vid in id_map.values():
        if vid not in tags_map:
            tags_map[vid] = []

    # Zusammenfassung
    vouchers_with_tags = sum(1 for tags in tags_map.values() if tags)
    print(f"[info] Found {vouchers_with_tags} vouchers with existing tags (via TagRelation)")

    if debug:
        for vid, tags in tags_map.items():
            if tags:
                print(f"[debug] Voucher {vid} has tags: {tags}")

    return tags_map

def format_currency(amount, currency='EUR'):
    """Formatiert Betrag mit Währung"""
    return f"{amount:.2f} {currency}"

def sanitize_filename(filename):
    """Bereinigt Dateinamen von ungültigen Zeichen"""
    # Ersetze ungültige Zeichen durch Unterstrich
    invalid_chars = r'[<>:"/\\|?*\x00-\x1f]'
    sanitized = re.sub(invalid_chars, '_', filename)
    # Entferne führende/nachfolgende Leerzeichen und Punkte
    sanitized = sanitized.strip(' .')
    # Begrenze Länge auf 200 Zeichen (vor Erweiterung)
    if len(sanitized) > 200:
        sanitized = sanitized[:200]
    return sanitized if sanitized else "unnamed"

def tag_vouchers_in_sevdesk(session, vouchers, tag_name, expected_credit_debit, verify_ssl=True, debug=False):
    """Erstellt Tags für Vouchers in sevDesk.
    
    BUG-FIX v0.1.30: Validiert creditDebit vor dem Tagging!
    
    Args:
        session: requests Session mit Auth-Header
        vouchers: Liste von Voucher-Daten (mit 'id' und 'creditDebit' Feld)
        tag_name: Name des Tags (z.B. "EXPORT_2025_Q4")
        expected_credit_debit: Erwarteter creditDebit-Wert ('D' für income, 'C' für expense)
        verify_ssl: SSL-Verifizierung
        debug: Debug-Modus
    
    Returns:
        tuple: (erfolgreiche_tags: int, fehlgeschlagene_tags: int, fehler_liste: list)
    """
    if not vouchers:
        return 0, 0, []
    
    url = "https://my.sevdesk.de/api/v1/Tag/Factory/create"
    successful = 0
    failed = 0
    skipped = 0
    errors = []
    total = len(vouchers)
    
    # ====== BUG-FIX v0.1.30: Debug-Ausgabe VOR dem Tagging ======
    print(f"\n[info] ========== TAGGING DEBUG (v0.1.30) ==========")
    print(f"[info] Anzahl zu taggender Vouchers: {total}")
    print(f"[info] Tag-Name: '{tag_name}'")
    print(f"[info] Erwarteter creditDebit: {expected_credit_debit}")
    
    # Zeige die ersten 3 Vouchers mit ihren creditDebit-Werten
    print(f"\n[info] Erste 3 Vouchers zur Validierung:")
    for i, v in enumerate(vouchers[:3], 1):
        v_id = v.get('id')
        v_nr = v.get('nr', 'N/A')
        v_cd = v.get('creditDebit', 'MISSING')
        print(f"       {i}. ID={v_id}, Nr={v_nr}, creditDebit={v_cd}")
    
    # Prüfe creditDebit-Verteilung in der zu taggenden Liste
    cd_distribution = {}
    for v in vouchers:
        cd = v.get('creditDebit', 'MISSING')
        cd_distribution[cd] = cd_distribution.get(cd, 0) + 1
    
    print(f"\n[info] creditDebit-Verteilung in Volltreffer-Liste:")
    for cd_val, count in sorted(cd_distribution.items()):
        if cd_val == 'D':
            label = "income/Ausgangsrechnungen"
        elif cd_val == 'C':
            label = "expense/Eingangsrechnungen"
        else:
            label = "MISSING/UNKNOWN"
        status = "✓ KORREKT" if cd_val == expected_credit_debit else "✗ FALSCH"
        print(f"       - {cd_val}: {count} vouchers ({label}) {status}")
    
    # Warnung bei falschen creditDebit-Werten
    wrong_type_count = sum(1 for v in vouchers if v.get('creditDebit') != expected_credit_debit)
    if wrong_type_count > 0:
        print(f"\n[WARNING] ================================================")
        print(f"[WARNING] {wrong_type_count} Vouchers haben falschen creditDebit-Wert!")
        print(f"[WARNING] Diese werden ÜBERSPRUNGEN und NICHT getaggt!")
        print(f"[WARNING] ================================================")
    
    print(f"[info] ================================================\n")
    
    print(f"[info] Starte Tagging für {total} Belege mit Tag '{tag_name}'...")
    
    for idx, voucher in enumerate(vouchers, 1):
        voucher_id = voucher.get('id')
        beleg_nr = voucher.get('nr', f"voucher_{voucher_id}")
        voucher_credit_debit = voucher.get('creditDebit', 'MISSING')
        
        # ====== BUG-FIX v0.1.30: Validiere creditDebit vor dem Tagging ======
        if voucher_credit_debit != expected_credit_debit:
            print(f"  Tagging {idx}/{total}: {beleg_nr} ... ⚠ ÜBERSPRUNGEN (creditDebit={voucher_credit_debit}, erwartet={expected_credit_debit})")
            skipped += 1
            errors.append({
                'nr': idx,
                'beleg_nr': beleg_nr,
                'voucher_id': voucher_id,
                'error': f"SKIPPED: Wrong creditDebit ({voucher_credit_debit} != {expected_credit_debit})"
            })
            continue
        
        print(f"  Tagging {idx}/{total}: {beleg_nr} ...", end=" ")
        
        payload = {
            "name": tag_name,
            "object": {
                "id": voucher_id,
                "objectName": "Voucher"
            }
        }
        
        if debug:
            print(f"\n[debug] Tag Request: {url}")
            print(f"[debug] Payload: {payload}")
        
        try:
            response = session.post(
                url, 
                json=payload, 
                verify=verify_ssl, 
                timeout=30
            )
            response.raise_for_status()
            print("✓")
            successful += 1
        except requests.exceptions.HTTPError as e:
            error_msg = f"HTTP {e.response.status_code}"
            try:
                error_detail = e.response.json()
                if 'error' in error_detail:
                    error_msg = error_detail.get('error', {}).get('message', error_msg)
            except:
                pass
            print(f"✗ ({error_msg})")
            failed += 1
            errors.append({
                'nr': idx,
                'beleg_nr': beleg_nr,
                'voucher_id': voucher_id,
                'error': error_msg
            })
        except requests.exceptions.Timeout:
            print("✗ (Timeout)")
            failed += 1
            errors.append({
                'nr': idx,
                'beleg_nr': beleg_nr,
                'voucher_id': voucher_id,
                'error': "Timeout"
            })
        except Exception as e:
            error_msg = str(e)
            print(f"✗ ({error_msg})")
            failed += 1
            errors.append({
                'nr': idx,
                'beleg_nr': beleg_nr,
                'voucher_id': voucher_id,
                'error': error_msg
            })
    
    # ====== BUG-FIX v0.1.30: Zeige Zusammenfassung nach dem Tagging ======
    print(f"\n[info] Tagging abgeschlossen:")
    print(f"       - Erfolgreich: {successful}")
    print(f"       - Fehlgeschlagen: {failed}")
    print(f"       - Übersprungen (falscher creditDebit): {skipped}")
    
    return successful, failed, errors

def download_voucher_pdf(session, voucher_id, output_path, verify_ssl=True, debug=False):
    """Lädt das PDF-Dokument eines Vouchers herunter.
    
    Endpoint: GET /Voucher/{voucherId}/downloadDocument
    
    Returns:
        tuple: (success: bool, error_message: str or None)
    """
    url = f"https://my.sevdesk.de/api/v1/Voucher/{voucher_id}/downloadDocument"
    
    if debug:
        print(f"[debug] PDF Download Request: {url}")
    
    try:
        response = session.get(url, verify=verify_ssl, timeout=60)
        response.raise_for_status()
        
        # Die API gibt JSON mit base64-encoded Content zurück
        data = response.json()
        
        if 'objects' in data and data['objects']:
            obj = data['objects']
            if isinstance(obj, dict):
                content = obj.get('content')
                base64_encoded = obj.get('base64Encoded', True)
            elif isinstance(obj, list) and len(obj) > 0:
                content = obj[0].get('content')
                base64_encoded = obj[0].get('base64Encoded', True)
            else:
                return False, "Unerwartetes Antwortformat"
            
            if content:
                if base64_encoded:
                    pdf_data = base64.b64decode(content)
                else:
                    pdf_data = content.encode('utf-8') if isinstance(content, str) else content
                
                with open(output_path, 'wb') as f:
                    f.write(pdf_data)
                return True, None
            else:
                return False, "Kein Content in der Antwort"
        else:
            # Möglicherweise direkter binary Response
            content_type = response.headers.get('Content-Type', '')
            if 'application/pdf' in content_type or 'image/' in content_type:
                with open(output_path, 'wb') as f:
                    f.write(response.content)
                return True, None
            return False, "Kein Dokument gefunden"
            
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            return False, "Dokument nicht gefunden (404)"
        return False, f"HTTP Fehler: {e.response.status_code}"
    except requests.exceptions.Timeout:
        return False, "Timeout beim Download"
    except Exception as e:
        return False, f"Fehler: {str(e)}"

def download_pdfs_for_vouchers(session, voucher_list, output_dir, verify_ssl=True, debug=False):
    """Lädt PDFs für eine Liste von Vouchers herunter.
    
    Args:
        session: requests Session mit Auth-Header
        voucher_list: Liste von Voucher-Daten (mit 'id' und 'nr' Feldern)
        output_dir: Zielverzeichnis für PDFs
        verify_ssl: SSL-Verifizierung
        debug: Debug-Modus
    
    Returns:
        tuple: (erfolgreiche_downloads: int, fehlgeschlagene_downloads: int, fehler_liste: list)
    """
    if not voucher_list:
        return 0, 0, []
    
    successful = 0
    failed = 0
    errors = []
    total = len(voucher_list)
    
    print(f"\n[info] Starte PDF-Download für {total} Belege...")
    
    for idx, voucher in enumerate(voucher_list, 1):
        voucher_id = voucher.get('id')
        beleg_nr = voucher.get('nr', f"voucher_{voucher_id}")
        
        # PDF-Dateiname: Laufende Nummer (idx) statt Beleg-Nr
        # idx entspricht der "Nr." Spalte in der XLSX-Datei
        pdf_path = os.path.join(output_dir, f"{idx}.pdf")
        
        print(f"  Downloading PDF {idx}/{total}: {beleg_nr} -> {idx}.pdf ...", end=" ")
        
        success, error = download_voucher_pdf(
            session, voucher_id, pdf_path, 
            verify_ssl=verify_ssl, debug=debug
        )
        
        if success:
            print("✓")
            successful += 1
        else:
            print(f"✗ ({error})")
            failed += 1
            errors.append({
                'nr': idx,
                'beleg_nr': beleg_nr,
                'voucher_id': voucher_id,
                'error': error
            })
    
    return successful, failed, errors

def create_xlsx_sheet(ws, data_list, voucher_type, is_income):
    """Erstellt ein formatiertes Sheet mit den Voucher-Daten.
    
    Spaltenreihenfolge: Nr., Rechn-Dat, Lief-Dat, Beleg-Nr, Lieferant/Kunde, Beschreibung, Betrag, Bereits exportiert (Tag)
    """
    # Header definieren
    lieferant_kunde = "Kunde" if is_income else "Lieferant"
    headers = ["Nr.", "Rechn-Dat", "Lief-Dat", "Beleg-Nr", lieferant_kunde, "Beschreibung", "Betrag", "Bereits exportiert (Tag)"]
    
    # Header-Styling
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header schreiben
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Daten schreiben
    for row_idx, item in enumerate(data_list, 2):
        nr = row_idx - 1  # Laufende Nummer beginnend bei 1
        ws.cell(row=row_idx, column=1, value=nr).border = thin_border
        ws.cell(row=row_idx, column=2, value=item['v_date']).border = thin_border
        ws.cell(row=row_idx, column=3, value=item['d_date']).border = thin_border
        ws.cell(row=row_idx, column=4, value=item['nr']).border = thin_border
        ws.cell(row=row_idx, column=5, value=item['supplier_customer']).border = thin_border
        ws.cell(row=row_idx, column=6, value=item['desc']).border = thin_border
        ws.cell(row=row_idx, column=7, value=item['amount']).border = thin_border
        
        # Rechtsbündige Ausrichtung für Betrag
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='right')
        
        # Neue Spalte: Bereits exportiert (Tag)
        existing_tags = item.get('existing_tags', [])
        tags_str = ", ".join(existing_tags) if existing_tags else "-"
        ws.cell(row=row_idx, column=8, value=tags_str).border = thin_border
    
    # Spaltenbreiten anpassen (neue Spalte hinzugefügt)
    column_widths = [6, 12, 12, 18, 30, 40, 15, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def create_output_directory(export_tag, voucher_type):
    """Erstellt die Verzeichnisstruktur und löscht nur das relevante Unterverzeichnis.
    
    Struktur:
    <exportTag>/
    ├── Ausgangsrechnungen/
    └── Eingangsrechnungen/
    
    Args:
        export_tag: Export-Tag-Name (z.B. "EXPORT_2025_Q4")
        voucher_type: 'income' oder 'expense'
    """
    # Hauptverzeichnis-Name ist jetzt der exportTag
    main_dir = export_tag
    
    # Unterverzeichnis basierend auf Typ bestimmen
    if voucher_type == 'income':
        target_subdir = "Ausgangsrechnungen"
    else:
        target_subdir = "Eingangsrechnungen"
    
    target_path = os.path.join(main_dir, target_subdir)
    
    # Hauptverzeichnis erstellen falls nicht vorhanden
    if not os.path.exists(main_dir):
        os.makedirs(main_dir, exist_ok=True)
        print(f"[info] Hauptverzeichnis erstellt: {main_dir}/")
    
    # NUR das relevante Unterverzeichnis löschen (falls vorhanden)
    if os.path.exists(target_path):
        print(f"[info] Lösche existierendes Unterverzeichnis: {target_path}")
        shutil.rmtree(target_path)
    
    # Unterverzeichnis erstellen
    os.makedirs(target_path, exist_ok=True)
    print(f"[info] Unterverzeichnis erstellt: {target_path}/")
    
    return main_dir

def main():
    parser = argparse.ArgumentParser(description=f"sevDesk Extrakt v{__version__}")
    parser.add_argument("--begin", required=True, help="Start Leistungszeitraum (DD.MM.YYYY oder YYYY-MM-DD)")
    parser.add_argument("--end", required=True, help="Ende Leistungszeitraum (DD.MM.YYYY oder YYYY-MM-DD)")
    parser.add_argument("--endRechnungsdatum", required=True, help="Stichtag Rechnungsdatum (DD.MM.YYYY oder YYYY-MM-DD)")
    parser.add_argument("--type", choices=['income', 'expense'], required=True)
    parser.add_argument("--exportTag", required=True, help="Export-Tag-Name für Verzeichnis und sevDesk-Tagging (z.B. 'EXPORT_2025_Q4', 'DEZEMBER_2025')")
    parser.add_argument("--token", help="sevDesk API Token")
    parser.add_argument("--no-verify", action="store_false", dest="verify_ssl", default=True)
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--skip-pdf", action="store_true", help="PDF-Download überspringen")
    parser.add_argument("--skip-tagging", action="store_true", help="sevDesk-Tagging überspringen")
    args = parser.parse_args()

    if not args.verify_ssl:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    api_token = args.token or os.environ.get("SEVDESK_API_TOKEN")
    session = requests.Session()
    session.headers.update({"Authorization": api_token})

    begin_dt = parse_date(args.begin)
    end_dt = parse_date(args.end)
    stichtag_dt = parse_date(args.endRechnungsdatum)

    # ====== Verbesserte Banner-Ausgabe ======
    print("=" * 60)
    print(f"sevDesk Extrakt v{__version__}")
    print("=" * 60)
    print(f"Requested type: {args.type}")
    print(f"  → API filter: creditDebit={'D' if args.type == 'income' else 'C'}")
    print(f"  → Expected: {'Ausgangsrechnungen (Debit)' if args.type == 'income' else 'Eingangsrechnungen (Credit)'}")
    print("=" * 60)

    # ====== BUG-FIX v0.1.29: Fetch vouchers mit verbessertem Logging ======
    vouchers = fetch_vouchers(
        session, 
        voucher_type=args.type,
        verify_ssl=args.verify_ssl, 
        debug=args.debug
    )
    
    # ====== LOKALE FILTERUNG als Sicherheitsnetz ======
    expected_credit_debit = 'D' if args.type == 'income' else 'C'
    original_count = len(vouchers)
    
    # Filtere Vouchers mit falschem creditDebit LOKAL heraus
    vouchers = [v for v in vouchers if v.get('creditDebit') == expected_credit_debit]
    filtered_out = original_count - len(vouchers)
    
    if filtered_out > 0:
        print(f"\n[WARNING] ================================================")
        print(f"[WARNING] Filtered out {filtered_out} vouchers with wrong creditDebit!")
        print(f"[WARNING] Original: {original_count}, After filter: {len(vouchers)}")
        print(f"[WARNING] This indicates the API filter is not working correctly!")
        print(f"[WARNING] ================================================")
    
    print(f"\n[info] Processing {len(vouchers)} vouchers of type '{args.type}' (creditDebit={expected_credit_debit})")
    
    # ====== Tags für die (bereits gefilterten) Vouchers abrufen ======
    tags_map = fetch_tags_for_all_vouchers(
        session=session,
        vouchers=vouchers,
        verify_ssl=args.verify_ssl,
        debug=args.debug
    )
    
    # 3-stufige Filterung: Listen für Volltreffer und Abgelehnt
    volltreffer_list = []
    abgelehnt_list = []
    
    # Zähler für bereits getaggte Vouchers
    already_tagged_count = 0

    is_income = (args.type == 'income')

    for v in vouchers:
        v_date_str = v.get('voucherDate')
        d_date_str = v.get('deliveryDate')
        if not v_date_str: continue

        try:
            v_dt = datetime.fromisoformat(v_date_str.replace('Z', '+00:00')).replace(tzinfo=None)
            d_dt = datetime.fromisoformat(d_date_str.replace('Z', '+00:00')).replace(tzinfo=None) if d_date_str else v_dt
        except: continue

        try:
            gross = float(v.get('sumGross', 0))
        except:
            gross = 0

        # ====== STUFE 1: Rechnungsdatum zwischen BEGIN und ENDRECHNUNGSDATUM ======
        if not (begin_dt <= v_dt <= stichtag_dt):
            continue

        # Voucher-Daten vorbereiten (inkl. ID für PDF-Download)
        supplier_customer = v.get('supplierName') or ''
        if not supplier_customer:
            supplier_obj = v.get('supplier')
            if supplier_obj and isinstance(supplier_obj, dict):
                supplier_customer = supplier_obj.get('name') or ''
                if not supplier_customer:
                    surename = supplier_obj.get('surename') or ''
                    familyname = supplier_obj.get('familyname') or ''
                    supplier_customer = f"{surename} {familyname}".strip()
        supplier_customer = supplier_customer or "Unbekannt"
        
        # Tags für diesen Voucher abrufen
        voucher_id = v.get('id')
        existing_tags = tags_map.get(voucher_id, [])
        
        # ====== BUG-FIX v0.1.30: creditDebit im voucher_data speichern! ======
        voucher_data = {
            "id": voucher_id,
            "creditDebit": v.get('creditDebit'),  # BUG-FIX: creditDebit speichern!
            "v_date": v_dt.strftime('%d.%m.%Y'),
            "d_date": d_dt.strftime('%d.%m.%Y'),
            "v_dt": v_dt,
            "nr": v.get('voucherNumber') or v.get('id'),
            "supplier_customer": supplier_customer,
            "desc": v.get('description') or "Kein Betreff",
            "amount": format_currency(gross, v.get('currency', 'EUR')),
            "existing_tags": existing_tags
        }

        # ====== STUFE 2: Lieferdatum zwischen BEGIN und END? ======
        lieferdatum_ok = (begin_dt <= d_dt <= end_dt)
        
        # ====== STUFE 3: Hat der Voucher bereits einen Tag? ======
        has_existing_tags = len(existing_tags) > 0
        
        if lieferdatum_ok:
            if has_existing_tags:
                abgelehnt_list.append(voucher_data)
                already_tagged_count += 1
            else:
                volltreffer_list.append(voucher_data)
        else:
            abgelehnt_list.append(voucher_data)

    # Sortieren nach Rechnungsdatum (absteigend)
    volltreffer_list.sort(key=lambda x: x['v_dt'], reverse=True)
    abgelehnt_list.sort(key=lambda x: x['v_dt'], reverse=True)

    # XLSX-Datei erstellen
    wb = Workbook()
    
    ws_volltreffer = wb.active
    ws_volltreffer.title = "Volltreffer"
    create_xlsx_sheet(ws_volltreffer, volltreffer_list, args.type, is_income)
    
    ws_abgelehnt = wb.create_sheet(title="Abgelehnt")
    create_xlsx_sheet(ws_abgelehnt, abgelehnt_list, args.type, is_income)

    # Dateinamen-Format
    begin_file_str = begin_dt.strftime('%d-%m-%Y')
    end_file_str = end_dt.strftime('%d-%m-%Y')
    
    # Verzeichnisstruktur erstellen
    main_dir = create_output_directory(args.exportTag, args.type)
    
    if args.type == 'income':
        subdir = "Ausgangsrechnungen"
        filename = f"ausgangsrechnungen_{begin_file_str}_{end_file_str}.xlsx"
    else:
        subdir = "Eingangsrechnungen"
        filename = f"eingangsrechnungen_{begin_file_str}_{end_file_str}.xlsx"
    
    output_dir = os.path.join(main_dir, subdir)
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    
    # ====== PDF-DOWNLOAD ======
    pdf_successful = 0
    pdf_failed = 0
    pdf_errors = []
    
    if not args.skip_pdf:
        if volltreffer_list:
            pdf_successful, pdf_failed, pdf_errors = download_pdfs_for_vouchers(
                session=session,
                voucher_list=volltreffer_list,
                output_dir=output_dir,
                verify_ssl=args.verify_ssl,
                debug=args.debug
            )
    
    # ====== SEVDESK TAGGING (BUG-FIX v0.1.30) ======
    tag_successful = 0
    tag_failed = 0
    tag_errors = []
    
    if not args.skip_tagging:
        if volltreffer_list:
            # BUG-FIX v0.1.30: Übergebe expected_credit_debit zur Validierung!
            tag_successful, tag_failed, tag_errors = tag_vouchers_in_sevdesk(
                session=session,
                vouchers=volltreffer_list,
                tag_name=args.exportTag,
                expected_credit_debit=expected_credit_debit,  # BUG-FIX: Neuer Parameter!
                verify_ssl=args.verify_ssl,
                debug=args.debug
            )
    
    # Zusammenfassung
    print(f"\n{'=' * 60}")
    print(f"ZUSAMMENFASSUNG - sevDesk Extrakt v{__version__}")
    print(f"{'=' * 60}")
    print(f"Typ: {'Ausgangsrechnungen (income/D)' if args.type == 'income' else 'Eingangsrechnungen (expense/C)'}")
    print(f"Export-Tag: {args.exportTag}")
    print(f"Zeitraum: {args.begin} - {args.end}")
    print(f"Rechnungsdatum bis: {args.endRechnungsdatum}")
    print("-" * 60)
    print(f"Vouchers von API abgerufen: {original_count}")
    if filtered_out > 0:
        print(f"  → davon {filtered_out} mit falschem creditDebit (lokal gefiltert)")
    print(f"Nach lokaler Filterung: {len(vouchers)}")
    print("-" * 60)
    print(f"Volltreffer: {len(volltreffer_list)}")
    print(f"Abgelehnt:   {len(abgelehnt_list)}")
    
    if already_tagged_count > 0:
        print(f"  → davon {already_tagged_count} bereits exportiert (getaggt)")
    
    print("-" * 60)
    print(f"XLSX gespeichert: {filepath}")
    
    if not args.skip_pdf:
        print("-" * 60)
        print(f"PDF-Downloads (nur Volltreffer):")
        print(f"  Gesamt: {len(volltreffer_list)}")
        print(f"  Erfolgreich: {pdf_successful}")
        print(f"  Fehlgeschlagen: {pdf_failed}")
        
        if pdf_errors:
            print(f"\nFehlgeschlagene Downloads:")
            for err in pdf_errors[:10]:
                print(f"  - Nr. {err['nr']} ({err['beleg_nr']}): {err['error']}")
            if len(pdf_errors) > 10:
                print(f"  ... und {len(pdf_errors) - 10} weitere Fehler")
    
    if not args.skip_tagging:
        print("-" * 60)
        print(f"sevDesk-Tagging (nur Volltreffer):")
        print(f"  Tag: '{args.exportTag}'")
        print(f"  Gesamt: {len(volltreffer_list)}")
        print(f"  Erfolgreich: {tag_successful}")
        print(f"  Fehlgeschlagen: {tag_failed}")
        
        if tag_errors:
            print(f"\nFehlgeschlagene Tags:")
            for err in tag_errors[:10]:
                print(f"  - Nr. {err['nr']} ({err['beleg_nr']}): {err['error']}")
            if len(tag_errors) > 10:
                print(f"  ... und {len(tag_errors) - 10} weitere Fehler")
    
    print("=" * 60)

if __name__ == "__main__":
    main()
